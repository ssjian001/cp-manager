"""Tests for Excel Export — 4 test cases covering export_control_plan()."""

from __future__ import annotations

import os

import pytest
from openpyxl import load_workbook

import db.database as db
import services.project_service as ps
import services.plan_service as pls
import services.item_service as its
import services.approval_service as asrv
from export.excel_export import export_control_plan


# ── Fixtures ──────────────────────────────────────────────────────────────


@pytest.fixture
def db_path(tmp_path, monkeypatch):
    """Redirect DB to tmp_path so tests don't touch ~/.cp-manager/."""
    test_db = tmp_path / "test.db"
    monkeypatch.setattr(db, "DB_PATH", test_db)
    monkeypatch.setattr(db, "DB_DIR", tmp_path)
    db.init_db()
    yield test_db


@pytest.fixture
def project(db_path):
    """Create a fully populated project."""
    pid = ps.create_project(
        name="Export Test Project",
        part_number="EX-001",
        part_name="Export Widget",
        supplier="Export Corp",
        supplier_code="EXC",
        contact_person="Jane Smith",
        contact_phone="555-0100",
    )
    return ps.get_project(pid)


@pytest.fixture
def full_plan(db_path, project):
    """Create a plan with steps, items, team members, and approvals."""
    pid = pls.create_plan(
        project["id"],
        cp_number="CP-EXPORT-01",
        phase="prototype",
        core_team="Alice, Bob",
    )

    # Team members
    asrv.add_team_member(pid, "Alice", role="Engineer", department="QE")
    asrv.add_team_member(pid, "Bob", role="Manager", department="Prod")

    # Approvals
    aid1 = asrv.create_approval(pid, "prepared", "Alice")
    aid2 = asrv.create_approval(pid, "approved", "Bob")
    asrv.sign_approval(aid1)
    asrv.sign_approval(aid2)

    # Steps
    conn = db.get_connection()
    try:
        cur = conn.execute(
            "INSERT INTO process_steps (plan_id, step_number, step_name, equipment, sort_order) "
            "VALUES (?, ?, ?, ?, ?)",
            (pid, "OP10", "Machining", "CNC", 0),
        )
        step1_id = cur.lastrowid
        cur = conn.execute(
            "INSERT INTO process_steps (plan_id, step_number, step_name, sort_order) "
            "VALUES (?, ?, ?, ?)",
            (pid, "OP20", "Inspection", 1),
        )
        step2_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    # Items — 2 for step1, 1 for step2 → 3 total items
    its.create_item(
        step1_id, pid,
        char_number="CH-001",
        char_type="product",
        char_description="Diameter",
        special_classification="CC",
        specification="10.0",
        tolerance="0.5",
        measurement_method="Caliper",
        gauge_id="G-001",
        sample_size="5",
        sample_frequency="1/hr",
        control_method_type="SPC",
        reaction_plan="停线通知挑选",
        responsible="Alice",
        ep_verification_freq="",
        ep_verification_method="",
    )
    its.create_item(
        step1_id, pid,
        char_number="CH-002",
        char_type="product",
        char_description="Length",
        special_classification="none",
        specification="100.0",
        tolerance="1.0",
        measurement_method="Micrometer",
        gauge_id="G-002",
        sample_size="3",
        sample_frequency="每批",
        control_method_type="manual",
        reaction_plan="停线通知",
        responsible="Bob",
    )
    its.create_item(
        step2_id, pid,
        char_number="CH-003",
        char_type="process",
        char_description="Temperature",
        special_classification="SC",
        specification="200°C",
        tolerance="5°C",
        measurement_method="Thermometer",
        gauge_id="G-003",
        sample_size="1",
        sample_frequency="1/班",
        control_method_type="EP",
        reaction_plan="停线隔离",
        ep_verification_freq="每日",
        ep_verification_method="Visual check",
    )

    return pls.get_plan(pid)


# ═══════════════════════════════════════════════════════════════════════════
# Tests
# ═══════════════════════════════════════════════════════════════════════════


class TestExcelExport:
    def test_export_creates_file(self, db_path, full_plan, tmp_path):
        """创建完整数据，导出，验证文件存在且大小 > 0"""
        output = tmp_path / "test_export.xlsx"
        result_path = export_control_plan(full_plan["id"], str(output))

        assert os.path.exists(result_path)
        assert os.path.getsize(result_path) > 0

    def test_export_has_correct_sheets(self, db_path, full_plan, tmp_path):
        """验证 sheet 名为 'Control Plan'"""
        output = tmp_path / "test_sheets.xlsx"
        export_control_plan(full_plan["id"], str(output))

        wb = load_workbook(str(output))
        sheet_names = wb.sheetnames
        assert "Control Plan" in sheet_names
        assert len(sheet_names) == 1  # only one sheet
        wb.close()

    def test_export_header_data(self, db_path, full_plan, project, tmp_path):
        """验证表头行包含正确的项目信息"""
        output = tmp_path / "test_header.xlsx"
        export_control_plan(full_plan["id"], str(output))

        wb = load_workbook(str(output))
        ws = wb.active

        # Row 1: 控制计划编号 / 零件号
        assert ws.cell(1, 1).value == "控制计划编号"
        assert ws.cell(1, 2).value == "CP-EXPORT-01"  # cp_number
        assert ws.cell(1, 7).value == "零件号/最新变更级别"
        assert ws.cell(1, 8).value == "EX-001"  # part_number

        # Row 2: 零件名称/描述 / 供应商/工厂
        assert ws.cell(2, 1).value == "零件名称/描述"
        assert ws.cell(2, 2).value == "Export Widget"  # part_name
        assert ws.cell(2, 7).value == "供应商/工厂"
        assert ws.cell(2, 8).value == "Export Corp"  # supplier

        # Row 3: 供应商代码 / 关键联系人/电话
        assert ws.cell(3, 1).value == "供应商代码"
        assert ws.cell(3, 2).value == "EXC"  # supplier_code
        assert ws.cell(3, 7).value == "关键联系人/电话"
        assert "Jane Smith" in ws.cell(3, 8).value
        assert "555-0100" in ws.cell(3, 8).value

        # Row 4: 核心团队
        assert ws.cell(4, 1).value == "日期(编制)"
        assert ws.cell(4, 5).value == "核心团队"
        assert ws.cell(4, 6).value == "Alice, Bob"

        # Row 5: approvals
        assert ws.cell(5, 1).value == "供应商批准/日期"
        assert "Alice" in (ws.cell(5, 2).value or "")

        # Row 7: phase markers
        assert "☑" in (ws.cell(7, 1).value or "")
        assert "Prototype" in (ws.cell(7, 1).value or "")

        # Column headers (row 8)
        assert ws.cell(8, 1).value == "过程编号"
        assert ws.cell(8, 2).value == "过程名称/操作描述"
        assert ws.cell(8, 6).value == "特殊特性分类"

        wb.close()

    def test_export_body_data(self, db_path, full_plan, tmp_path):
        """验证表体行数等于控制项数"""
        output = tmp_path / "test_body.xlsx"
        export_control_plan(full_plan["id"], str(output))

        wb = load_workbook(str(output))
        ws = wb.active

        # Header is rows 1-8, body starts at row 9
        # We have 3 items total (2 in OP10, 1 in OP20)
        # Each item is one row → 3 body rows
        # Find last non-empty row
        max_row = ws.max_row
        body_rows = max_row - 8  # rows 9..N

        assert body_rows == 3, f"Expected 3 body rows, got {body_rows}"

        # Check first item's data
        assert ws.cell(9, 1).value == "OP10"  # step number
        assert ws.cell(9, 2).value == "Machining"  # step name
        assert ws.cell(9, 4).value == "CH-001"  # first char number
        assert ws.cell(9, 6).value == "CC"  # special classification

        # Check second item (same step, row 10)
        assert ws.cell(10, 4).value == "CH-002"

        # Check third item (OP20, row 11)
        assert ws.cell(11, 1).value == "OP20"
        assert ws.cell(11, 4).value == "CH-003"
        assert ws.cell(11, 6).value == "SC"

        wb.close()
