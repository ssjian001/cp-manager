"""AIAG 标准控制计划 Excel 导出。

基于 AIAG Control Plan 1st Edition (2024) 格式，使用 openpyxl 生成。
"""

import os
from datetime import datetime

import db.database as db
import services.plan_service as plan_svc
import services.item_service as item_svc
import services.approval_service as approval_svc
import services.project_service as project_svc

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    msg = (
        "openpyxl is required for Excel export. "
        "Install it with: uv add openpyxl  (or pip install openpyxl)"
    )
    raise ImportError(msg)


# ── Constants ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", bold=True, size=10)

EVEN_ROW_FILL = PatternFill(
    start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
)
ODD_ROW_FILL = PatternFill(
    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
)
SPECIAL_FILL = PatternFill(
    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
)
SAFE_LAUNCH_FILL = PatternFill(
    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
)

THIN_SIDE = Side(style="thin", color="000000")
CELL_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
CENTER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="left")

# Column widths (approximate)
COL_WIDTHS = {
    "A": 10,   # 过程编号
    "B": 18,   # 过程名称/操作描述
    "C": 18,   # 机器/装置/夹具/工具
    "D": 10,   # 特性编号
    "E": 22,   # 产品/过程特性描述
    "F": 10,   # 特殊特性分类
    "G": 20,   # 产品/过程规格/公差
    "H": 18,   # 评价/测量技术
    "I": 8,    # 样本量
    "J": 12,   # 样本频率
    "K": 12,   # 控制方法
    "L": 18,   # 反应计划
    "M": 12,   # RESP
    "N": 12,   # EP/MP验证频次
    "O": 12,   # EP/MP验证方法
}

# Column headers (row 8)
COLUMN_HEADERS = [
    "过程编号",
    "过程名称/操作描述",
    "机器/装置/夹具/工具",
    "特性编号",
    "产品/过程特性描述",
    "特殊特性分类",
    "产品/过程规格/公差",
    "评价/测量技术",
    "样本量",
    "样本频率",
    "控制方法",
    "反应计划",
    "RESP",
    "EP/MP验证频次",
    "EP/MP验证方法",
]


# ── Helpers ─────────────────────────────────────────────────────────────────────


def _apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    """Apply styling to a single cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def _set_header_cell(ws, row, col, value, merge_end=None):
    """Write a header-area cell with merged-blue-header styling."""
    cell = ws.cell(row=row, column=col, value=value)
    _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN, border=CELL_BORDER)
    if merge_end:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_end,
        )
        # Apply style to all merged cells
        for c in range(col + 1, merge_end + 1):
            mc = ws.cell(row=row, column=c)
            _apply_cell_style(mc, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN, border=CELL_BORDER)


def _set_data_cell(ws, row, col, value, is_special=False, is_safe_launch=False, is_header=False):
    """Write a body-region cell with appropriate styling."""
    cell = ws.cell(row=row, column=col, value=value or "")

    if is_header:
        # Column header row (row 8)
        _apply_cell_style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN, border=CELL_BORDER)
    else:
        if is_special:
            fill = SPECIAL_FILL
        elif is_safe_launch:
            fill = SAFE_LAUNCH_FILL
        else:
            fill = EVEN_ROW_FILL if (row % 2 == 0) else ODD_ROW_FILL

        align = CENTER_ALIGN if col in (1, 4, 6, 9, 10) else LEFT_ALIGN
        _apply_cell_style(cell, font=BODY_FONT, fill=fill, alignment=align, border=CELL_BORDER)

    return cell


def _format_datetime(dt_str: str | None) -> str:
    """Format a datetime string to YYYY-MM-DD."""
    if not dt_str:
        return ""
    try:
        dt = datetime.fromisoformat(dt_str)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return str(dt_str)[:10]


def _get_approval_text(approvals: list[dict], approval_type: str) -> str:
    """Get formatted approval text: 'name/date'."""
    for a in approvals:
        if a["approval_type"] == approval_type:
            name = a.get("name", "") or ""
            date = _format_datetime(a.get("signed_at"))
            if name and date:
                return f"{name}/{date}"
            return name or date or ""
    return ""


def _get_phase_markers(phase: str, is_safe_launch: bool) -> str:
    """Build the phase markers line (row 7)."""
    markers = []
    for p, label in [("prototype", "Prototype"), ("pre_launch", "Pre-Launch"), ("production", "Production")]:
        checked = "☑" if phase == p else "☐"
        markers.append(f"{checked} {label}")

    sl_label = "Safe Launch"
    sl_checked = "☑" if is_safe_launch else "☐"
    markers.append(f"{sl_checked} {sl_label}")

    return "    ".join(markers)


# ── Main Export Function ───────────────────────────────────────────────────────


def export_control_plan(plan_id: int, output_path: str) -> str:
    """导出控制计划到 Excel。返回文件路径。"""
    # ── 1. Fetch data ──────────────────────────────────────────────────────
    plan = plan_svc.get_plan(plan_id)
    if not plan:
        raise ValueError(f"Control plan {plan_id} not found")

    # Fetch project
    project = project_svc.get_project(plan["project_id"])

    # Fetch process steps
    conn = db.get_connection()
    try:
        steps_raw = conn.execute(
            "SELECT * FROM process_steps WHERE plan_id = ? ORDER BY sort_order",
            (plan_id,),
        ).fetchall()
    finally:
        conn.close()
    steps = [dict(r) for r in steps_raw]

    # Fetch items
    items = item_svc.list_items(plan_id)

    # Fetch approvals
    approvals = approval_svc.list_approvals(plan_id)

    # ── 2. Build lookup: step_id -> list of items ──────────────────────────
    items_by_step: dict[int, list[dict]] = {}
    for item in items:
        items_by_step.setdefault(item["step_id"], []).append(item)

    # ── 3. Create workbook ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Plan"

    # ── 4. Set column widths ───────────────────────────────────────────────
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── 5. Render header (rows 1-7) ────────────────────────────────────────
    # Row 1: 控制计划编号 / 零件号
    _set_header_cell(ws, 1, 1, "控制计划编号")
    _set_header_cell(ws, 1, 2, plan.get("cp_number") or "", merge_end=6)
    _set_header_cell(ws, 1, 7, "零件号/最新变更级别")
    _set_header_cell(ws, 1, 8, (project or {}).get("part_number") or "", merge_end=12)

    # Row 2: 零件名称/描述 / 供应商/工厂
    _set_header_cell(ws, 2, 1, "零件名称/描述")
    _set_header_cell(ws, 2, 2, (project or {}).get("part_name") or "", merge_end=6)
    _set_header_cell(ws, 2, 7, "供应商/工厂")
    _set_header_cell(ws, 2, 8, (project or {}).get("supplier") or "", merge_end=12)

    # Row 3: 供应商代码 / 关键联系人/电话
    _set_header_cell(ws, 3, 1, "供应商代码")
    _set_header_cell(ws, 3, 2, (project or {}).get("supplier_code") or "", merge_end=6)
    _set_header_cell(ws, 3, 7, "关键联系人/电话")
    contact_parts = []
    cp = (project or {}).get("contact_person") or ""
    ph = (project or {}).get("contact_phone") or ""
    if cp:
        contact_parts.append(cp)
    if ph:
        contact_parts.append(ph)
    _set_header_cell(ws, 3, 8, " / ".join(contact_parts), merge_end=12)

    # Row 4: 日期(编制) / 日期(修订) / 核心团队
    _set_header_cell(ws, 4, 1, "日期(编制)")
    _set_header_cell(ws, 4, 2, _format_datetime(plan.get("created_at")))
    _set_header_cell(ws, 4, 3, "日期(修订)")
    _set_header_cell(ws, 4, 4, _format_datetime(plan.get("updated_at")))
    _set_header_cell(ws, 4, 5, "核心团队")
    _set_header_cell(ws, 4, 6, plan.get("core_team") or "", merge_end=12)

    # Row 5: 供应商批准/日期 / 客户工程批准/日期
    prepared_text = _get_approval_text(approvals, "prepared")
    approved_text = _get_approval_text(approvals, "approved")
    _set_header_cell(ws, 5, 1, "供应商批准/日期")
    _set_header_cell(ws, 5, 2, prepared_text, merge_end=6)
    _set_header_cell(ws, 5, 7, "客户工程批准/日期")
    _set_header_cell(ws, 5, 8, approved_text, merge_end=12)

    # Row 6: empty
    for col in range(1, 16):
        c = ws.cell(row=6, column=col, value="")
        _apply_cell_style(c, border=CELL_BORDER)

    # Row 7: phase markers
    phase_markers = _get_phase_markers(
        plan.get("phase", "prototype"),
        bool(plan.get("is_safe_launch")),
    )
    _set_header_cell(ws, 7, 1, phase_markers, merge_end=15)
    # Also set row height for phase marker row
    ws.row_dimensions[7].height = 30

    # ── 6. Render column headers (row 8) ───────────────────────────────────
    for col_idx, header_text in enumerate(COLUMN_HEADERS, start=1):
        _set_data_cell(ws, 8, col_idx, header_text, is_header=True)

    # ── 7. Render body data (row 9+) ───────────────────────────────────────
    current_row = 9

    for step in steps:
        step_items = items_by_step.get(step["id"], [])

        if not step_items:
            # Step with no items — write one row with step info
            _set_data_cell(ws, current_row, 1, step.get("step_number"))
            _set_data_cell(ws, current_row, 2, step.get("step_name"))
            _set_data_cell(ws, current_row, 3, step.get("equipment"))
            for col in range(4, 16):
                _set_data_cell(ws, current_row, col, "")
            current_row += 1
        else:
            step_start_row = current_row
            for item in step_items:
                is_special = item.get("special_classification", "none") not in (
                    "none",
                    "",
                )
                is_sl = bool(plan.get("is_safe_launch"))

                _set_data_cell(ws, current_row, 1, step.get("step_number"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 2, step.get("step_name"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 3, step.get("equipment"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 4, item.get("char_number"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 5, item.get("char_description"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 6, item.get("special_classification"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 7, _format_spec(item), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 8, _format_measurement(item), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 9, item.get("sample_size"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 10, item.get("sample_frequency"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 11, item.get("control_method_type"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 12, item.get("reaction_plan"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 13, item.get("responsible"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 14, item.get("ep_verification_freq"), is_special=is_special, is_safe_launch=is_sl)
                _set_data_cell(ws, current_row, 15, item.get("ep_verification_method"), is_special=is_special, is_safe_launch=is_sl)
                current_row += 1

            # Merge step columns (A, B, C) if more than one item in this step
            if len(step_items) > 1:
                step_end_row = current_row - 1
                if step_end_row > step_start_row:
                    for merge_col in (1, 2, 3):
                        ws.merge_cells(
                            start_row=step_start_row,
                            start_column=merge_col,
                            end_row=step_end_row,
                            end_column=merge_col,
                        )

    # ── 8. Set print area and page setup ────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = "A9"

    # ── 9. Save ─────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    return os.path.abspath(output_path)


# ── Internal Helpers ───────────────────────────────────────────────────────────


def _format_spec(item: dict) -> str:
    """Combine specification and tolerance into a single string."""
    spec = (item.get("specification") or "").strip()
    tol = (item.get("tolerance") or "").strip()
    if spec and tol:
        return f"{spec} ± {tol}"
    return spec or tol or ""


def _format_measurement(item: dict) -> str:
    """Combine measurement_method and gauge_id into a single string."""
    method = (item.get("measurement_method") or "").strip()
    gauge = (item.get("gauge_id") or "").strip()
    if method and gauge:
        return f"{method} [{gauge}]"
    return method or gauge or ""
